import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import json
import re
from datetime import datetime

CONFIG_FILE = "analyzer_settings.json"

class ProductionAnalyzerAppTrueFinal:
    def __init__(self, master):
        self.master = master
        self.master.title("지능형 생산 분석 시스템 (v3.3 - DB 최적화)")
        self.master.geometry("850x850")

        self.production_df, self.capacity_df, self.criteria_df, self.defect_df = None, None, None, None
        self.target_dfs = {}
        self.available_target_dates = []

        self.prod_file_path = ""
        self.yield_settings, self.util_settings, self.target_settings, self.defect_settings = {}, {}, {}, {}

        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.pack(fill="both", expand=True)

        file_frame = ttk.LabelFrame(main_frame, text="1. 파일 선택 (결과 파일 제외 자동 로딩)")
        file_frame.pack(fill="x", padx=5, pady=5)
        self.setup_file_loader(file_frame, "생산 실적 파일:", 0, self.load_production_file)
        self.prod_file_label = self.last_label
        self.setup_file_loader(file_frame, "최대 생산량 파일:", 1, self.load_capacity_file)
        self.cap_file_label = self.last_label
        self.setup_file_loader(file_frame, "연도별 생산 목표 파일:", 2, self.load_target_file)
        self.target_file_label = self.last_label
        self.setup_file_loader(file_frame, "저가동 기준 파일:", 3, self.load_criteria_file)
        self.criteria_file_label = self.last_label
        self.setup_file_loader(file_frame, "불량 실적 파일:", 4, self.load_defect_file)
        self.defect_file_label = self.last_label

        date_frame = ttk.LabelFrame(main_frame, text="2. 기간 필터링 (비워두면 전체 기간)")
        date_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(date_frame, text="시작일:").grid(row=0, column=0, padx=5, pady=5, sticky="w"); self.start_date_entry = ttk.Entry(date_frame); self.start_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Label(date_frame, text="종료일:").grid(row=1, column=0, padx=5, pady=5, sticky="w"); self.end_date_entry = ttk.Entry(date_frame); self.end_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew"); date_frame.columnconfigure(1, weight=1)

        agg_frame = ttk.LabelFrame(main_frame, text="3. 데이터 내보내기 기준")
        agg_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(agg_frame, text="시간 단위:").pack(side="left", padx=5, pady=5)
        self.time_agg_var = tk.StringVar(value='일별')
        self.time_agg_combo = ttk.Combobox(agg_frame, textvariable=self.time_agg_var, values=['일별', '주간별', '월별', '연도별'], state="readonly")
        self.time_agg_combo.pack(side="left", padx=5, pady=5, fill="x", expand=True)

        mode_frame = ttk.LabelFrame(main_frame, text="4. 분석 모드 선택"); mode_frame.pack(fill="x", padx=5, pady=5); self.mode_var = tk.StringVar(value="수율 분석");
        modes = ["수율 분석", "가동률 분석", "목표 달성률 분석", "저가동 설비 분석", "불량 원인 분석"]
        [ttk.Radiobutton(mode_frame, text=mode, variable=self.mode_var, value=mode, command=self.on_mode_change).pack(side="left", padx=10, pady=5) for mode in modes]

        self.group_by_frame = ttk.LabelFrame(main_frame, text="5. 데이터 요약 기준"); self.group_by_frame.pack(fill="x", padx=5, pady=5); self.group_vars = {};
        group_options = ['생산일자', '공장', '공정코드', '신규분류요약', '함수율', '품명', '기계코드', '사출기계코드', '공정기계코드', '불량명']
        [self.group_vars.update({option: tk.BooleanVar()}) or ttk.Checkbutton(self.group_by_frame, text=option, variable=self.group_vars[option]).grid(row=i//5, column=i%5, padx=5, pady=5, sticky='w') for i, option in enumerate(group_options)]

        action_frame = ttk.Frame(main_frame); action_frame.pack(fill="x", padx=5, pady=20); self.generate_button = ttk.Button(action_frame, text="보고서 생성", command=self.generate_report); self.generate_button.pack(pady=5, fill="x", ipady=5)

        self.status_bar = ttk.Label(self.master, text="준비 완료", relief="sunken", anchor="w", padding=5); self.status_bar.pack(side="bottom", fill="x")

        self.current_mode = self.mode_var.get()
        self.load_settings()
        self.auto_load_default_files()
        self.on_mode_change(is_initial_call=True)
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _save_multisheet_excel_autofit(self, sheets_data, file_path):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                history_col_name = '과거 생산 품목 상세 이력'

                for i, column_name in enumerate(df.columns, 1):
                    column_letter = get_column_letter(i)
                    if column_name == history_col_name:
                        worksheet.column_dimensions[column_letter].width = 99
                        for cell in worksheet[column_letter]:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        try:
                            max_length = max((df[column_name].astype(str).map(len).max(), len(str(column_name))))
                            worksheet.column_dimensions[column_letter].width = max_length + 4
                        except (ValueError, TypeError):
                            worksheet.column_dimensions[column_letter].width = len(str(column_name)) + 4

    def _save_df_to_excel_autofit(self, df, file_path):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Summary')
            worksheet = writer.sheets['Summary']
            history_col_name = '과거 생산 품목 상세 이력'

            for i, column_name in enumerate(df.columns, 1):
                column_letter = get_column_letter(i)
                if column_name == history_col_name:
                    worksheet.column_dimensions[column_letter].width = 99
                    for cell in worksheet[column_letter]:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    try:
                        max_length = max((df[column_name].astype(str).map(len).max(), len(str(column_name))))
                        worksheet.column_dimensions[column_letter].width = max_length + 2
                    except (ValueError, TypeError):
                        worksheet.column_dimensions[column_letter].width = len(str(column_name)) + 2

    def generate_defect_report(self):
        self.status_bar.config(text="불량 원인 분석 보고서 생성 중...")
        self.master.update()
        if self.defect_df is None:
            messagebox.showwarning("경고", "'불량 실적 파일'을 선택해야 합니다.")
            return

        group_by_columns = [col for col, var in self.group_vars.items() if var.get()]
        if not group_by_columns:
            messagebox.showwarning("경고", "집계 기준을 하나 이상 선택해주세요. (예: 공장, 불량명)")
            return

        try:
            df = self.defect_df.copy()

            original_cols = list(df.columns)
            found_defect_cols = [col for col in original_cols if str(col).startswith('불량수량')]

            if len(found_defect_cols) < 2:
                messagebox.showerror("파일 구조 오류", f"'불량실적현황' 파일에 '불량수량'으로 시작하는 컬럼이 2개 이상 필요합니다.\n(현재 {len(found_defect_cols)}개 발견됨)")
                return
                
            df.rename(columns={
                found_defect_cols[0]: '불량수량(전체)',
                found_defect_cols[1]: '불량수량(유형별)'
            }, inplace=True)

            df['생산일자'] = pd.to_datetime(df['생산일자'].astype(str).str.replace('.', '-'), errors='coerce')
            df.dropna(subset=['생산일자'], inplace=True)
            start_date_str = self.start_date_entry.get().replace('.', '-')
            end_date_str = self.end_date_entry.get().replace('.', '-')
            if start_date_str: df = df[df['생산일자'] >= pd.to_datetime(start_date_str)]
            if end_date_str: df = df[df['생산일자'] <= pd.to_datetime(end_date_str)]

            if df.empty:
                messagebox.showinfo("정보", "선택된 기간에 해당하는 데이터가 없습니다.")
                return

            numeric_cols = ['양품수량', '불량수량(전체)', '불량수량(유형별)']
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

            prod_runs_df = df.drop_duplicates(subset=['생산실적번호'])
            prod_group_cols = [col for col in group_by_columns if col != '불량명' and col in df.columns]
            
            if prod_group_cols:
                prod_agg_df = prod_runs_df.groupby(prod_group_cols).agg(
                    양품수량=('양품수량', 'sum'),
                    불량수량_전체_집계=('불량수량(전체)', 'sum')
                ).reset_index()
            else: # 사용자가 '불량명'만 선택하는 등, 생산량을 묶을 기준이 없는 경우
                prod_agg_df = pd.DataFrame([{
                    '양품수량': prod_runs_df['양품수량'].sum(),
                    '불량수량_전체_집계': prod_runs_df['불량수량(전체)'].sum()
                }])
            
            prod_agg_df['생산수량'] = prod_agg_df['양품수량'] + prod_agg_df['불량수량_전체_집계']

            detail_group_cols = [col for col in group_by_columns if col in df.columns]
            if not detail_group_cols:
                 messagebox.showwarning("경고", "'데이터 요약 기준'에서 유효한 컬럼을 선택하세요.")
                 return
                 
            defect_agg_df = df.groupby(detail_group_cols).agg(
                불량수량_유형별_집계=('불량수량(유형별)', 'sum')
            ).reset_index()

            if prod_group_cols:
                final_df = pd.merge(defect_agg_df, prod_agg_df, on=prod_group_cols, how='left')
            else:
                final_df = defect_agg_df.assign(**prod_agg_df.iloc[0])

            final_df['불량률(%)'] = (final_df['불량수량_전체_집계'] / final_df['생산수량'] * 100).where(final_df['생산수량'] > 0, 0)
            final_df['분석일시'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            final_df.rename(columns={'불량수량_전체_집계': '불량수량(전체)', '불량수량_유형별_집계': '불량수량(유형별)'}, inplace=True)
            
            final_cols_order = detail_group_cols + [
                '생산수량', '양품수량', '불량수량(전체)', '불량수량(유형별)', '불량률(%)', '분석일시'
            ]
            final_df = final_df[[col for col in final_cols_order if col in final_df.columns]]
            final_df = final_df.sort_values(by=detail_group_cols, ascending=True)

            save_path = "불량실적현황(최적화).xlsx"
            sheets_to_save = {"설비별_상세분석": final_df}
            self._save_multisheet_excel_autofit(sheets_to_save, save_path)
            messagebox.showinfo("성공", f"최적화된 불량 분석 보고서가 생성되었습니다.\n위치: {save_path}")
            self.status_bar.config(text="불량 원인 분석 완료.")

        except Exception as e:
            import traceback
            messagebox.showerror("오류", f"보고서 생성 중 예측하지 못한 오류 발생: {e}\n\n{traceback.format_exc()}")
            self.status_bar.config(text="오류 발생")


    def generate_low_utilization_report(self):
        self.status_bar.config(text="저가동 설비 분석 중..."); self.master.update()
        if self.production_df is None or self.capacity_df is None or self.criteria_df is None:
            messagebox.showwarning("경고", "분석을 위해 '생산 실적', '최대 생산량', '저가동 기준' 파일이 모두 필요합니다.")
            return

        try:
            prod_df = self.production_df.copy()
            prod_df['생산일자'] = pd.to_datetime(prod_df['생산일자'].astype(str).str.replace('.', '-'), errors='coerce')
            prod_df.dropna(subset=['생산일자'], inplace=True)

            start_date_str = self.start_date_entry.get().replace('.', '-')
            end_date_str = self.end_date_entry.get().replace('.', '-')
            start_date = pd.to_datetime(start_date_str) if start_date_str else prod_df['생산일자'].min()
            end_date = pd.to_datetime(end_date_str) if end_date_str else prod_df['생산일자'].max()
            all_dates = pd.date_range(start=start_date, end=end_date, freq='D')

            machine_keys = ['공장', '공정코드', '기계코드']
            all_machines = self.criteria_df[machine_keys].drop_duplicates()
            if all_machines.empty:
                messagebox.showwarning("경고", "'저가동 기준 파일'에 분석할 설비 정보가 없습니다."); return

            scaffold_index = pd.MultiIndex.from_product([all_machines.to_records(index=False), all_dates], names=['machine', '생산일자'])
            scaffold_df = pd.DataFrame(scaffold_index.tolist(), columns=['machine','생산일자'])
            scaffold_df[machine_keys] = pd.DataFrame(scaffold_df['machine'].tolist(), index=scaffold_df.index)
            scaffold_df.drop('machine', axis=1, inplace=True)

            prod_df['생산수량'] = pd.to_numeric(prod_df['생산수량'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
            daily_prod_summary = prod_df.groupby(machine_keys + ['생산일자'])['생산수량'].sum().reset_index()

            daily_util_df = pd.merge(scaffold_df, daily_prod_summary, on=machine_keys + ['생산일자'], how='left')
            daily_util_df['생산수량'].fillna(0, inplace=True)

            daily_util_df = pd.merge(daily_util_df, self.capacity_df, on=machine_keys, how='left')
            daily_util_df['이론상 최대 생산량'].fillna(0, inplace=True)

            daily_util_df['일별 가동률(%)'] = (daily_util_df['생산수량'] / daily_util_df['이론상 최대 생산량'].where(daily_util_df['이론상 최대 생산량'] != 0)) * 100
            daily_util_df['일별 가동률(%)'].fillna(0, inplace=True)

            avg_util_summary = daily_util_df.groupby(machine_keys)['일별 가동률(%)'].mean().reset_index()
            report_with_criteria = pd.merge(avg_util_summary, self.criteria_df, on=machine_keys, how='left')
            report_with_criteria.dropna(subset=['저가동설비기준'], inplace=True)
            low_util_machines = report_with_criteria[report_with_criteria['일별 가동률(%)'] <= report_with_criteria['저가동설비기준']].copy()
            low_util_machines.rename(columns={'일별 가동률(%)': '기간 내 가동률(%)'}, inplace=True)

            if low_util_machines.empty:
                messagebox.showinfo("정보", "지정된 기간에 기준 미달인 저가동 설비가 없습니다."); return

            full_prod_df = self.production_df.copy(); [full_prod_df.__setitem__(col, 'N/A') for col in ['품명', '신규분류요약', '함수율'] if col not in full_prod_df.columns]; full_prod_df.fillna({'품명': '', '신규분류요약': '', '함수율': ''}, inplace=True)
            history_df = full_prod_df[full_prod_df['기계코드'].isin(low_util_machines['기계코드'].unique())].copy()
            history_df.drop_duplicates(subset=['기계코드', '품명', '신규분류요약', '함수율'], inplace=True)

            def format_history(df_group):
                output_parts = [f"분류: {category}, 함수율: {moisture}\n  - 품명: {', '.join(sorted(group['품명'].unique()))}" for (category, moisture), group in df_group.groupby(['신규분류요약', '함수율'])]
                return "\n\n".join(output_parts)
            prod_history = history_df.groupby('기계코드').apply(format_history).reset_index(name='과거 생산 품목 상세 이력')

            final_report_df = pd.merge(low_util_machines, prod_history, on='기계코드', how='left')
            final_report_df['저가동설비기준'] = final_report_df['저가동설비기준'].round(2).astype(str) + '%'
            final_report_df['기간 내 가동률(%)'] = final_report_df['기간 내 가동률(%)'].round(2).astype(str) + '%'
            final_report_df = final_report_df[['공장', '공정코드', '기계코드', '저가동설비기준', '기간 내 가동률(%)', '과거 생산 품목 상세 이력']]
            final_report_df.fillna({'과거 생산 품목 상세 이력': '이력 없음'}, inplace=True)

            save_path = f"{os.path.splitext(self.prod_file_path)[0]}(저가동설비).xlsx"
            self._save_df_to_excel_autofit(final_report_df, save_path)
            messagebox.showinfo("성공", f"저가동 설비 분석 보고서가 생성되었습니다.\n위치: {save_path}")
            self.status_bar.config(text="저가동 설비 분석 완료.")

        except Exception as e:
            messagebox.showerror("오류", f"보고서 생성 중 오류 발생: {e}"); self.status_bar.config(text="오류 발생")

    def _load_file(self, file_path, file_type):
        label_widget_map = {'prod': self.prod_file_label, 'capa': self.cap_file_label, 'target': self.target_file_label, 'criteria': self.criteria_file_label, 'defect': self.defect_file_label}
        success_text_map = {'prod': "생산 실적 로딩 완료", 'capa': "최대 생산량 로딩 완료", 'target': "월별 생산 목표 로딩 완료", 'criteria': "저가동 기준 로딩 완료", 'defect': "불량 실적 로딩 완료"}
        label_widget, success_text = label_widget_map.get(file_type), success_text_map.get(file_type)
        if file_type == 'prod': self.prod_file_path = file_path
        self.status_bar.config(text=f"'{os.path.basename(file_path)}' 읽는 중..."); self.master.update()
        try:
            df_dict = pd.read_excel(file_path, sheet_name=None)

            if file_type == 'target':
                self.target_dfs.clear()
                all_sheets_df = pd.concat(df_dict.values(), ignore_index=True)

                required_cols = ['년', '월', '공장', '공정코드', '일일_생산목표량']
                if not all(col in all_sheets_df.columns for col in required_cols):
                    raise ValueError(f"생산 목표 파일에는 {', '.join(required_cols)} 컬럼이 모두 필요합니다.")

                all_sheets_df.dropna(subset=required_cols, inplace=True)
                all_sheets_df['년'] = pd.to_numeric(all_sheets_df['년'], errors='coerce').astype('Int64')
                all_sheets_df['월'] = pd.to_numeric(all_sheets_df['월'], errors='coerce').astype('Int64')
                all_sheets_df.dropna(subset=['년', '월'], inplace=True)
                
                for (year, month), group in all_sheets_df.groupby(['년', '월']):
                    self.target_dfs[(year, month)] = group
                
                self.available_target_dates = sorted(self.target_dfs.keys())
                
                if self.available_target_dates:
                    min_date = f"{self.available_target_dates[0][0]}년 {self.available_target_dates[0][1]}월"
                    max_date = f"{self.available_target_dates[-1][0]}년 {self.available_target_dates[-1][1]}월"
                    success_text = f"월별 목표 로딩 완료: {min_date} ~ {max_date}"
                else:
                    success_text = "월별 목표 데이터 없음"

                setattr(self, 'target_df_loaded', True)

            elif file_type == 'criteria':
                df = pd.concat(df_dict.values(), ignore_index=True)
                criteria_col_name = '저가동설비기준'
                if criteria_col_name not in df.columns: raise KeyError(f"'{criteria_col_name}' 컬럼을 찾을 수 없습니다. 엑셀 파일의 D열 첫 행에 컬럼명이 올바르게 입력되었는지 확인해주세요.")
                s = df[criteria_col_name]
                if pd.api.types.is_numeric_dtype(s) and (s.dropna() <= 1).all() and (s.dropna() > 0).any(): df[criteria_col_name] = s * 100
                else: df[criteria_col_name] = pd.to_numeric(s.astype(str).str.replace('%', '', regex=False), errors='coerce')
                df.dropna(subset=[criteria_col_name], inplace=True)
                self.criteria_df = df
            else:
                df_attribute_map = {'prod': 'production_df', 'capa': 'capacity_df', 'defect': 'defect_df'}
                df_attribute = df_attribute_map.get(file_type)
                sheets_to_concat = list(df_dict.keys())
                if not sheets_to_concat: raise ValueError("유효한 시트를 찾을 수 없습니다.")
                df = pd.concat([df_dict[s] for s in sheets_to_concat], ignore_index=True)
                setattr(self, df_attribute, df)

            label_widget.config(text=os.path.basename(file_path), foreground="black"); self.status_bar.config(text=success_text)
        except Exception as e:
            messagebox.showerror("오류", f"'{os.path.basename(file_path)}' 파일 읽기 오류: {e}")
            label_widget.config(text="파일 없음", foreground="gray")
            if file_type == 'criteria': self.criteria_df = None
            if file_type == 'defect': self.defect_df = None

    def on_mode_change(self, is_initial_call=False):
        settings_map = {"수율 분석": self.yield_settings, "가동률 분석": self.util_settings,
                        "목표 달성률 분석": self.target_settings, "불량 원인 분석": self.defect_settings}
        
        if not is_initial_call:
            settings_to_save = settings_map.get(self.current_mode)
            if settings_to_save is not None:
                for col, var in self.group_vars.items():
                    settings_to_save[col] = var.get()

        new_mode = self.mode_var.get()
        self.group_by_frame.config(text=f"5. 데이터 요약 기준 (현재 모드: {new_mode})")
        self.generate_button.config(text=f"{new_mode} 생성")

        if new_mode in ["저가동 설비 분석"]:
            self.group_by_frame.pack_forget()
            self.time_agg_combo.config(state="disabled")
        else:
            self.group_by_frame.pack(fill="x", padx=5, pady=5)
            self.time_agg_combo.config(state="readonly" if new_mode != "불량 원인 분석" else "disabled")
            if new_mode == "불량 원인 분석": self.time_agg_combo.set('일별')

            settings_to_load = settings_map.get(new_mode, {})
            for col, var in self.group_vars.items():
                var.set(settings_to_load.get(col, False))

        self.current_mode = new_mode

    def auto_load_default_files(self):
        keyword_map = {"생산실적현황": 'prod', "가동율참고": 'capa', "생산목표량": 'target', "설비리스트및저가동설비기준": 'criteria', "불량실적현황": 'defect'}
        exclude_suffixes = ["(수율)", "(가동률)", "(목표달성율)", "(저가동설비)", "(최적화)"]
        for keyword, file_type in keyword_map.items():
            found_file = next((f for f in os.listdir('.') if keyword in f and f.endswith('.xlsx') and not any(suffix in f for suffix in exclude_suffixes)), None)
            if found_file: self._load_file(found_file, file_type)

    def setup_file_loader(self, parent, text, row, command):
        ttk.Label(parent, text=text).grid(row=row, column=0, padx=5, pady=5, sticky="w"); self.last_label = ttk.Label(parent, text="파일 없음", width=70, foreground="gray"); self.last_label.grid(row=row, column=1, padx=5, pady=5); ttk.Button(parent, text="수동 선택", command=command).grid(row=row, column=2, padx=5, pady=5)

    def load_criteria_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]);
        if path: self._load_file(path, 'criteria')
        
    def load_defect_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]);
        if path: self._load_file(path, 'defect')

    def get_settings_by_mode(self, mode):
        return {'수율 분석': self.yield_settings, '가동률 분석': self.util_settings, '목표 달성률 분석': self.target_settings, '불량 원인 분석': self.defect_settings}.get(mode)

    def on_closing(self):
        active_settings = self.get_settings_by_mode(self.current_mode);
        if active_settings is not None: [active_settings.update({col: var.get()}) for col, var in self.group_vars.items()]
        settings = {"yield_settings": self.yield_settings, "util_settings": self.util_settings, "target_settings": self.target_settings, "defect_settings": self.defect_settings};
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f: json.dump(settings, f, indent=4, ensure_ascii=False)
        self.master.destroy()

    def load_settings(self):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                self.yield_settings = settings.get("yield_settings", {})
                self.util_settings = settings.get("util_settings", {})
                self.target_settings = settings.get("target_settings", {})
                self.defect_settings = settings.get("defect_settings", {})
            self.status_bar.config(text="이전 설정을 불러왔습니다.")
        except FileNotFoundError:
            self.status_bar.config(text="초기 설정입니다.");
            self.util_settings['기계코드'] = True
            self.target_settings.update({'공장': True, '공정코드': True})
            self.defect_settings.update({'공장': True, '사출기계코드': True, '공정기계코드': True, '불량명': True})

    def load_production_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]);
        if path: self._load_file(path, 'prod')

    def load_capacity_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]);
        if path: self._load_file(path, 'capa')

    def load_target_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]);
        if path: self._load_file(path, 'target')

    def _prepare_base_df(self):
        if self.production_df is None: messagebox.showwarning("경고", "생산 실적 파일을 선택해주세요."); return None
        df = self.production_df.copy(); numeric_cols = ['양품수량', '불량수량', '샘플수량', '생산수량']; [df.__setitem__(col, pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)) for col in numeric_cols]
        df['생산일자'] = pd.to_datetime(df['생산일자'].astype(str).str.replace('.', '-'), errors='coerce'); start_date = self.start_date_entry.get().replace('.', '-'); end_date = self.end_date_entry.get().replace('.', '-')
        if start_date: df = df[df['생산일자'] >= pd.to_datetime(start_date)]
        if end_date: df = df[df['생산일자'] <= pd.to_datetime(end_date)]
        return df

    def _apply_time_aggregation(self, df, group_by_cols):
        if '생산일자' in group_by_cols:
            time_agg_unit = self.time_agg_var.get()
            dt_series = df['생산일자'].dt
            group_by_cols.remove('생산일자')
            if time_agg_unit == '일별':
                df['기간'] = dt_series.strftime('%Y-%m-%d')
                group_by_cols.insert(0, '기간')
            elif time_agg_unit == '주간별':
                week_periods = dt_series.to_period('W')
                df['기간'] = week_periods.apply(lambda p: f"{p.start_time.strftime('%Y-%m-%d')} ~ {p.end_time.strftime('%Y-%m-%d')}")
                df['주차'] = week_periods.apply(lambda p: f"{p.start_time.month}월 {(p.start_time.day - 1) // 7 + 1}주차")
                group_by_cols.insert(0, '주차')
                group_by_cols.insert(0, '기간')
            elif time_agg_unit == '월별':
                df['기간'] = dt_series.to_period('M').astype(str)
                group_by_cols.insert(0, '기간')
            elif time_agg_unit == '연도별':
                df['기간'] = dt_series.year
                group_by_cols.insert(0, '기간')
        return df, group_by_cols

    def generate_report(self):
        mode = self.mode_var.get()
        handler = {'수율 분석': self.generate_yield_report,
                   '가동률 분석': self.generate_utilization_report,
                   '목표 달성률 분석': self.generate_target_report,
                   '저가동 설비 분석': self.generate_low_utilization_report,
                   '불량 원인 분석': self.generate_defect_report}.get(mode)
        if handler: handler()

    def generate_yield_report(self):
        self.status_bar.config(text="수율 보고서 생성 중..."); self.master.update(); base_df = self._prepare_base_df()
        if base_df is None: return
        group_by_columns = [col for col, var in self.group_vars.items() if var.get()]
        if not group_by_columns: messagebox.showwarning("경고", "집계 기준을 선택해주세요."); return
        try:
            base_df, group_by_columns = self._apply_time_aggregation(base_df, group_by_columns)
            summary = base_df.groupby(group_by_columns).agg(총_생산수량=('생산수량', 'sum'), 총_양품수량=('양품수량', 'sum'), 총_불량수량=('불량수량', 'sum')).reset_index()
            summary['전체_수율(%)'] = round((summary['총_양품수량'] / summary['총_생산수량'].where(summary['총_생산수량'] != 0)) * 100, 2).fillna(0)
            save_path = f"{os.path.splitext(self.prod_file_path)[0]}(수율).xlsx"
            self._save_df_to_excel_autofit(summary, save_path); messagebox.showinfo("성공", f"수율 보고서가 생성되었습니다.\n위치: {save_path}"); self.status_bar.config(text="수율 보고서 생성 완료.")
        except Exception as e: messagebox.showerror("오류", f"보고서 생성 중 오류 발생: {e}")

    def generate_utilization_report(self):
        self.status_bar.config(text="가동률 보고서 생성 중..."); self.master.update();
        if self.capacity_df is None: messagebox.showwarning("경고", "'최대 생산량 파일'을 선택해야 합니다."); return
        base_df = self._prepare_base_df();
        if base_df is None: return
        group_by_columns = [col for col, var in self.group_vars.items() if var.get()]
        if not group_by_columns: messagebox.showwarning("경고", "집계 기준을 선택해주세요."); return
        try:
            base_df, group_by_columns = self._apply_time_aggregation(base_df, group_by_columns)
            merged_df = pd.merge(base_df, self.capacity_df, on=['공장', '공정코드', '기계코드'], how='left'); merged_df['이론상 최대 생산량'].fillna(0, inplace=True)
            agg_dict = {'총_생산수량': ('생산수량', 'sum'), '총_양품수량': ('양품수량', 'sum'), '일일_최대생산량': ('이론상 최대 생산량', 'first'), '운영일수': ('생산일자', 'nunique')}
            summary = merged_df.groupby(group_by_columns).agg(**agg_dict).reset_index()
            if self.time_agg_var.get() == '주간별': summary['운영일수'] = 7
            summary['이론상_총_생산량'] = summary['일일_최대생산량'] * summary['운영일수']
            summary['전체_수율(%)'] = round((summary['총_양품수량'] / summary['총_생산수량'].where(summary['총_생산수량'] != 0)) * 100, 2).fillna(0)
            summary['가동률(%)'] = round((summary['총_생산수량'] / summary['이론상_총_생산량'].where(summary['이론상_총_생산량'] != 0)) * 100, 2).fillna(0)
            final_cols = group_by_columns + ['총_생산수량', '총_양품수량', '전체_수율(%)', '운영일수', '이론상_총_생산량', '가동률(%)']
            summary = summary[[col for col in final_cols if col in summary.columns]]
            save_path = f"{os.path.splitext(self.prod_file_path)[0]}(가동률).xlsx"
            self._save_df_to_excel_autofit(summary, save_path); messagebox.showinfo("성공", f"가동률 보고서가 생성되었습니다.\n위치: {save_path}"); self.status_bar.config(text="가동률 보고서 생성 완료.")
        except Exception as e: messagebox.showerror("오류", f"보고서 생성 중 오류 발생: {e}")

    def _find_closest_target_df(self, year, month):
        target_date = (year, month)
        if target_date in self.target_dfs:
            return self.target_dfs[target_date]

        # Find the closest past date
        past_dates = [d for d in self.available_target_dates if d < target_date]
        if past_dates:
            return self.target_dfs[max(past_dates)]

        # Find the closest future date
        future_dates = [d for d in self.available_target_dates if d > target_date]
        if future_dates:
            return self.target_dfs[min(future_dates)]

        return None

    def generate_target_report(self):
        self.status_bar.config(text="목표 달성률 보고서 생성 중..."); self.master.update()
        if not hasattr(self, 'target_df_loaded') or not self.target_dfs:
            messagebox.showwarning("경고", "'월별 생산 목표 파일'을 선택해야 합니다.")
            return
        base_df = self._prepare_base_df()
        if base_df is None: return
        group_by_columns = [col for col, var in self.group_vars.items() if var.get()]
        if not group_by_columns:
            messagebox.showwarning("경고", "집계 기준을 선택해주세요.")
            return
        try:
            base_df['연도'] = base_df['생산일자'].dt.year
            base_df['월'] = base_df['생산일자'].dt.month

            processed_dfs = []
            for (prod_year, prod_month), group in base_df.groupby(['연도', '월']):
                target_df = self._find_closest_target_df(prod_year, prod_month)
                if target_df is not None:
                    merged_month_df = pd.merge(group, target_df, on=['공장', '공정코드'], how='left')
                    processed_dfs.append(merged_month_df)

            if not processed_dfs:
                messagebox.showinfo("정보", "선택된 기간에 해당하는 생산 목표 데이터가 없습니다.")
                return

            merged_df = pd.concat(processed_dfs, ignore_index=True)
            merged_df.dropna(subset=['일일_생산목표량'], inplace=True)
            merged_df = merged_df[merged_df['일일_생산목표량'] > 0]

            if merged_df.empty:
                messagebox.showinfo("정보", "유효한 생산 목표가 설정된 공정이 없습니다.")
                return

            merged_df, group_by_columns = self._apply_time_aggregation(merged_df, group_by_columns)

            agg_dict = {
                '총_생산수량': ('생산수량', 'sum'),
                '총_양품수량': ('양품수량', 'sum'),
                '일일_목표량': ('일일_생산목표량', 'first'),
                '운영일수': ('생산일자', 'nunique')
            }
            summary = merged_df.groupby(group_by_columns).agg(**agg_dict).reset_index()

            if self.time_agg_var.get() == '주간별': summary['운영일수'] = 7
            summary['목표_총_생산량'] = summary['일일_목표량'] * summary['운영일수']
            
            summary['양품수_기준_달성률(%)'] = round((summary['총_양품수량'] / summary['목표_총_생산량'].where(summary['목표_총_생산량'] != 0)) * 100, 2).fillna(0)
            
            final_cols_order = group_by_columns + [
                '총_양품수량', '총_생산수량', '목표_총_생산량', '양품수_기준_달성률(%)',
                '일일_목표량', '운영일수'
            ]
            summary = summary[[col for col in final_cols_order if col in summary.columns]]

            save_path = f"{os.path.splitext(self.prod_file_path)[0]}(목표달성율).xlsx"
            self._save_df_to_excel_autofit(summary, save_path)
            messagebox.showinfo("성공", f"목표 달성률 보고서가 생성되었습니다.\n위치: {save_path}")
            self.status_bar.config(text="목표 달성률 보고서 생성 완료.")
        except Exception as e:
            messagebox.showerror("오류", f"보고서 생성 중 오류 발생: {e}")

if __name__ == '__main__':
    root = tk.Tk()
    app = ProductionAnalyzerAppTrueFinal(root)
    root.mainloop()